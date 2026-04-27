from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from copy import deepcopy
import argparse


device_config = {
    "j784s4": {
        "rails_details": {
            'vdd_mcu_0v85': {
                'out_v': 0.85,
                'eff_ratio': 1,
                'group': "SOC_MCU"
            },
            'vdd_mcu_ram_0v85': {
                'out_v': 0.85,
                'eff_ratio': 1,
                'group': "SOC_RAM"
            },
            'vda_mcu_1v8': {
                'out_v': 1.8,
                'eff_ratio': 1,
                'group': "SOC"
            },
            'vdd_mcuio_3v3': {
                'out_v': 3.3,
                'eff_ratio': 1,
                'group': "SOC"
            },
            'vdd_mcuio_1v8': {
                'out_v': 1.8,
                'eff_ratio': 1,
                'group': "SOC"
            },
            'vdd_ram_0v85': {
                'out_v': 0.85,
                'eff_ratio': 1,
                'group': "SOC_RAM"
            },
            'vdd_wk_0v8': {
                'out_v': 0.8,
                'eff_ratio': 1,
                'group': "SOC"
            },
            'vdd_mcu_gpioret_3v3': {
                'out_v': 3.3,
                'eff_ratio': 1,
                'group': "SOC"
            },
            'vdd_phycore_0v8': {
                'out_v': 0.8,
                'eff_ratio': 1,
                'group': "SOC"
            },
            'vda_pll_1v8': {
                'out_v': 1.8,
                'eff_ratio': 1,
                'group': "SOC"
            },
            'vdd_phy_1v8': {
                'out_v': 1.8,
                'eff_ratio': 1,
                'group': "SOC"
            },
            'vda_usb_3v3': {
                'out_v': 3.3,
                'eff_ratio': 1,
                'group': "SOC"
            },
            'vdd_gpioret_3v3': {
                'out_v': 3.3,
                'eff_ratio': 1,
                'group': "SOC"
            },
            'vdd_io_1v8': {
                'out_v': 1.8,
                'eff_ratio': 1,
                'group': "SOC"
            },
            'vdd_io_3v3': {
                'out_v': 3.3,
                'eff_ratio': 1,
                'group': "SOC"
            },
            'vdd_sd_dv': {
                'out_v': 1.8,
                'eff_ratio': 1,
                'group': "SOC"
            },
            'vdd1_ddr_1v8': {
                'out_v': 1.8,
                'eff_ratio': 1,
                'group': "LPDDR4_SDRAM"
            },
            'vcca_core_3v3': {
                'out_v': 0.8,
                'eff_ratio': 0.8,
                'group': "SOC_CORE"
            },
            'vsys_mcuio_1v8': {
                'out_v': 1.8,
                'eff_ratio': 1,
                'group': "PERIPHERALS"
            },
            'vsys_mcuio_3v3': {
                'out_v': 3.3,
                'eff_ratio': 1,
                'group': "PERIPHERALS"
            },
            'vsys_io_1v8': {
                'out_v': 1.8,
                'eff_ratio': 1,
                'group': "PERIPHERALS"
            },
            'vsys_io_3v3': {
                'out_v': 3.3,
                'eff_ratio': 1,
                'group': "PERIPHERALS"
            },
            'vcc_12v0': {
                'out_v': 12,
                'eff_ratio': 1,
                'group': "VIN"
            },
            'vsys_5v0': {
                'out_v': 5,
                'eff_ratio': 1,
                'group': "VIN"
            },
            'vsys_3v3': {
                'out_v': 3.3,
                'eff_ratio': 1,
                'group': "VIN"
            },
            'vcca_ddr_3v3': {
                'out_v': 1.1,
                'eff_ratio': 0.8,
                'group': "SOC_DDR"
            },
            'vda_dll_0v8': {
                'out_v': 0.8,
                'eff_ratio': 1,
                'group': "SOC"
            },
            'vcca_cpu_avs_3v3': {
                'out_v': 0.8,
                'eff_ratio': 0.8,
                'group': "SOC_CPU"
            },
        },
        "no_calc_required": ["vdd1_ddr_1v8", "vsys_mcuio_1v8", "vsys_mcuio_3v3", "vsys_io_1v8", "vsys_io_3v3", "vcc_12v0", "vsys_5v0", "vsys_3v3"],
        "custom_calculation": ["vcca_core_3v3", "vcca_ddr_3v3", "vcca_cpu_avs_3v3"]
    },
    "j722s": {
        "rails_details": {
            "vdd_core": {
                'out_v': 0.8,
                'eff_ratio': 1,
                'group': "SOC"
            },
            "vdd_ram_0v85": {
                'out_v': 0.85,
                'eff_ratio': 1,
                'group': "SOC_RAM"
            },
            "vda_phy_1v8": {
                'out_v': 1.8,
                'eff_ratio': 1,
                'group': "SOC"
            },
            "vdd_ioret_core": {
                'out_v': 0.8,
                'eff_ratio': 1,
                'group': "SOC"
            },
            "vdd_sd_dv": {
                'out_v': 3.3,
                'eff_ratio': 1,
                'group': "SOC"
            },
            "vdd_io_1v8": {
                'out_v': 1.8,
                'eff_ratio': 1,
                'group': "SOC"
            },
            "vsys_3v3": {
                'out_v': 3.3,
                'eff_ratio': 1,
                'group': "VIN"
            },
            "vdd_ddr_1v1": {
                'out_v': 1.1,
                'eff_ratio': 1,
                'group': "SOC_DDR"
            },
            "vdd_io_3v3": {
                'out_v': 3.3,
                'eff_ratio': 1,
                'group': "SOC"
            },
            "vda_pll_1v8": {
                'out_v': 1.8,
                'eff_ratio': 1,
                'group': "SOC"
            },
            "vdd1_ddr_1v8": {
                'out_v': 1.8,
                'eff_ratio': 1,
                'group': "LPDDR4_SDRAM"
            }
        },
        "no_calc_required": ["vsys_3v3", "vdd1_ddr_1v8"],
        "custom_calculation": []
    }
}

class PowerReportGenerator:
    def __init__(self, config):
        self.config = config
        self.power_reading_list = []
        self.processed_data = []
        self.sheet_data_list = []

    def process_power_readings(self, idle_active_power_list=None):
        """Processes power readings to structure them for report generation."""
        # adding data
        idle_active_power_list_copy = deepcopy(idle_active_power_list)
        for index, reading_list in enumerate(idle_active_power_list_copy):
            if index == 0:  # Idle power before demo run
                for i in range(1, len(reading_list[0])):
                    reading_list[0][i] = "Idle " + reading_list[0][i]
                self.power_reading_list.extend(reading_list)
            else:  # Active power while demo is running
                for i in range(1, len(reading_list[0])):
                    reading_list[0][i] = "Active " + reading_list[0][i]
                for i in range(len(reading_list)):
                    self.power_reading_list[i].extend(reading_list[i][1:])

        # Append additional configuration details
        self.add_config_details()

    def add_config_details(self):
        """Adds configuration details like groups, output voltage, and efficiency ratio."""
        print("Adding configuration details...")
        for row_index, row in enumerate(self.power_reading_list):
            if row_index == 0:
                row.extend(
                    ["Group", "Output Voltage", "Efficiency Ratio"]
                )
            elif row[0] in self.config["rails_details"]:
                rail_details = self.config["rails_details"][row[0]]
                row.extend([
                    rail_details["group"],
                    rail_details["out_v"],
                    rail_details["eff_ratio"],
                ])


    def calculate_actual_power(self):
        """Performs actual power calculations based on configurations."""
        print("Calculating actual power...")
        for row_index, row in enumerate(self.power_reading_list):
            if row_index == 0: # headers
                # short formula here
                row.extend(
                    ["Idle Actual Current (mA)", "Idle Actual Power (mW)", "Active Actual Current (mA)", "Active Actual Power (mW)", "Power Difference"]
                )
            else:
                n = row_index + 2 # as data row starts from 3 in xlsx
                if row[0] in self.config["no_calc_required"]:
                    row.extend(["", "", "", "", ""])
                elif row[0] in self.config["custom_calculation"]:
                    row.extend([
                        f"=ROUND(D{n} * C{n} / K{n} * L{n}, 2)",
                        f"=ROUND(M{n} * K{n}, 2)",
                        f"=ROUND(G{n} * H{n} / K{n} * L{n}, 2)",
                        f"=ROUND(O{n} * K{n}, 2)",
                        f"=ROUND(P{n} - N{n}, 2)",
                    ])
                else:
                    row.extend([
                        f"=D{n}",
                        f"=ROUND(D{n} * C{n}, 2)",
                        f"=H{n}",
                        f"=ROUND(H{n} * G{n}, 2)",
                        f"=ROUND(P{n} - N{n}, 2)",
                    ])
    
    def prepare_final_sheet_data(self):
        """Prepares the final sheet data for the report generation."""
        # Adding headers
        self.sheet_data_list.append(
            ["", "Idle State Measurement Before Demo Run", "", "", "", "Demo Running State Measurements", "", "", ""]
        )

        # adding data
        self.sheet_data_list.extend(self.power_reading_list)

        # Adding Footer
        row_numbers = len(self.power_reading_list) + 1
        self.sheet_data_list.extend(
            [
                ["Total", "", "", "", "", "", "", "", "", "", "", "", "", f"=SUM(N3:N{row_numbers})", "", f"=SUM(P3:P{row_numbers})", f"=SUM(Q3:Q{row_numbers})"],
                ["", "", "", "", "", "", "", "", "", "", "", "", "", "Idle State Power", "", "Active State Power", "Power Difference [mW]"]
            ]
        )
        

    def generate_excel_report(self, filename="Power_Report.xlsx"):
        """Generates an Excel report with proper styling."""
        wb = Workbook()
        ws = wb.active

        # Write data to the worksheet
        for row in self.sheet_data_list:
            ws.append(row)

        # Apply styles
        self.apply_styles(ws)

        # Save the workbook
        wb.save(filename)
        print("Excel report generated successfully.")

    def apply_styles(self, ws):
        print("Applying styles...")
        def apply_style(row_start, row_end, col_start, col_end, color=None, bold=False, center=False, border=False):
            """Apply styles to a range of cells."""
            # Iterate over all rows in the range.
            for row in range(row_start, row_end + 1):
                # Iterate over all columns in the range.
                for col in range(col_start, col_end + 1):
                    # Get the cell at the current row and column.
                    cell = ws[f"{get_column_letter(col)}{row}"]
                    # Apply the fill color.
                    if color:
                        cell.fill = PatternFill(start_color=color, fill_type="solid")
                    # Apply the font style.
                    if bold:
                        cell.font = Font(bold=True)
                    # Apply the alignment.
                    if center:
                        cell.alignment = Alignment(horizontal="center")
                    # Apply the border.
                    if border:
                        cell.border = Border(
                            left=Side(style="thin"),
                            right=Side(style="thin"),
                            top=Side(style="thin"),
                            bottom=Side(style="thin"),
                        )

        # Header styling
        apply_style(2, 2, 1, len(self.sheet_data_list[2]), color="00b0f0", bold=True, center=True) # 2nd row header style blue color

        apply_style(1, 2, 2, 5, color="92d050", bold=True, center=True) # 1st row header style (idle state) green color
        apply_style(1, 2, 6, 9, color="fabf8f", bold=True, center=True) # 1st row header style (active state) yellow color

        apply_style(len(self.sheet_data_list)-1, len(self.sheet_data_list)-1, 1, len(self.sheet_data_list[2]), color="bfe4ef", bold=True, center=True, border=True) # Footer style light blue
        apply_style(len(self.sheet_data_list), len(self.sheet_data_list), 1, len(self.sheet_data_list[2]), bold=True, center=True) # Footer style light blue

        apply_style(2, 2, 13, 13, color="92d050", bold=True, center=True) # 2nd row header style (idle current calculation) green color
        apply_style(2, len(self.sheet_data_list), 13, 13, center=True) # 2nd row header style (idle power calculation) green color
        apply_style(2, len(self.sheet_data_list), 14, 14, color="92d050", center=True, border=True) # N column style

        apply_style(2, 2, 15, 15, color="fabf8f", bold=True, center=True) # 2nd row header style (actual current calculation) yellow color
        apply_style(2, len(self.sheet_data_list), 15, 15, center=True) # 2nd row header style (actual power calculation) yellow color
        apply_style(2, len(self.sheet_data_list), 16, 16, color="fabf8f", center=True, border=True) # P column style

        apply_style(2, len(self.sheet_data_list), 17, 17, color="b1a0c7", center=True, border=True) # Q column style for Power Difference
        apply_style(len(self.sheet_data_list)-1, len(self.sheet_data_list)-1, 17, 17, color="b1a0c7", bold=True, center=True, border=True) # Q column style for Power Difference

        for row_index, row in enumerate(self.sheet_data_list):
            # Apply red color for no calculation rows
            if row[0] in self.config['no_calc_required']:
                apply_style(row_index + 1, row_index + 1, 1, len(self.sheet_data_list[row_index]), color="da9694")

        # Apply border to all cells
        for row in ws.iter_rows(min_row=1, max_row=len(self.sheet_data_list)-1, min_col=1, max_col=len(self.sheet_data_list[1])):
            for cell in row:
                cell.border = Border(
                    left=Side(style="thin"),
                    right=Side(style="thin"),
                    top=Side(style="thin"),
                    bottom=Side(style="thin"),
                )

        # merge calls in first row
        ws.merge_cells(f'B1:E1')
        ws.merge_cells(f'F1:I1')

def parse_power_reading_str(data_str):
    """Parses the power readings into a structured list."""
    data_list = []
    for line in data_str.split('\n'):
        row_data_arr = [i.strip() for i in line.split("|")]
        if len(row_data_arr) == 8:
            row_data_arr = row_data_arr[2:-1]
            if row_data_arr[0] not in ["---", "", "END", "Total"]:
                data_list.append(row_data_arr)
    return data_list

# Usage
if __name__ == "__main__":
    # Takeing command line args
    parser = argparse.ArgumentParser(
        description="Generate Power Report from Power Readings")
    parser.add_argument("--soc", type=str, required=True, help="SoC name (e.g., j784s4, j722s)")
    parser.add_argument("--input_txt_file", type=str, required=True, help="Path to the input power readings file with `=>` separated readings")
    parser.add_argument("--output_excel_file", type=str, required=True, help="Path to save the output Excel report")
    args = parser.parse_args()

    # Read power readings from file
    power_readings = ""
    with open(args.input_txt_file, "r") as f:
        power_readings = f.read()

    # Parsing data
    power_data_list = [parse_power_reading_str(reading) for reading in power_readings.split("=>")]
    valid_power_data_list = [ data for data in power_data_list if len(data) > 10 and len(data[0]) == 5]

    report_generator = PowerReportGenerator(device_config[args.soc])
    report_generator.process_power_readings(idle_active_power_list=[valid_power_data_list[0], valid_power_data_list[1]])
    report_generator.calculate_actual_power()
    report_generator.prepare_final_sheet_data()
    report_generator.generate_excel_report(args.output_excel_file)
